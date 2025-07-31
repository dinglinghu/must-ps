"""
航天领域专业甘特图Excel生成器
生成包含详细数据和图表的Excel文件
"""

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from .realistic_constellation_gantt import ConstellationGanttData
from .gantt_save_config_manager import get_gantt_save_config_manager

logger = logging.getLogger(__name__)

class AerospaceGanttExcelGenerator:
    """航天领域专业甘特图Excel生成器"""
    
    def __init__(self):
        self.config_manager = get_gantt_save_config_manager()
        
        # 威胁等级颜色（Excel格式）
        self.threat_colors = {
            5: 'FFDC2626',  # 红色
            4: 'FFEA580C',  # 橙红色
            3: 'FFD97706',  # 橙色
            2: 'FF65A30D',  # 黄绿色
            1: 'FF059669',  # 绿色
            0: 'FF6B7280'   # 灰色
        }
        
        # 状态颜色
        self.status_colors = {
            'planned': 'FF6B7280',
            'ready': 'FF3B82F6',
            'executing': 'FFF59E0B',
            'completed': 'FF10B981',
            'failed': 'FFEF4444',
            'cancelled': 'FF8B5CF6',
            'paused': 'FFF97316'
        }
        
        logger.info("✅ 航天甘特图Excel生成器初始化完成")
    
    def generate_excel_gantt(
        self,
        gantt_data: ConstellationGanttData,
        output_path: str
    ) -> str:
        """生成Excel格式甘特图"""
        try:
            logger.info(f"📊 开始生成Excel甘特图: {output_path}")
            
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 创建各个工作表
            self._create_summary_sheet(wb, gantt_data)
            self._create_tasks_sheet(wb, gantt_data)
            self._create_satellites_sheet(wb, gantt_data)
            self._create_timeline_sheet(wb, gantt_data)
            self._create_statistics_sheet(wb, gantt_data)
            
            # 保存文件
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            
            logger.info(f"✅ Excel甘特图已生成: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"❌ 生成Excel甘特图失败: {e}")
            raise
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, gantt_data: ConstellationGanttData):
        """创建概要工作表"""
        ws = wb.create_sheet("任务概要")
        
        # 设置标题
        ws['A1'] = "航天任务调度甘特图概要"
        ws['A1'].font = Font(size=16, bold=True, color='FF1F2937')
        ws.merge_cells('A1:D1')
        
        # 基本信息
        info_data = [
            ["任务场景", gantt_data.mission_scenario],
            ["图表类型", gantt_data.chart_type],
            ["图表ID", gantt_data.chart_id],
            ["创建时间", gantt_data.creation_time.strftime('%Y-%m-%d %H:%M:%S')],
            ["开始时间", gantt_data.start_time.strftime('%Y-%m-%d %H:%M:%S')],
            ["结束时间", gantt_data.end_time.strftime('%Y-%m-%d %H:%M:%S')],
            ["总时长", f"{(gantt_data.end_time - gantt_data.start_time).total_seconds() / 3600:.2f} 小时"],
            ["任务数量", len(gantt_data.tasks)],
            ["卫星数量", len(gantt_data.satellites)],
            ["目标数量", len(gantt_data.missiles)]
        ]
        
        for i, (key, value) in enumerate(info_data, start=3):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # 性能指标
        if gantt_data.performance_metrics:
            ws['A15'] = "性能指标"
            ws['A15'].font = Font(size=14, bold=True)
            
            row = 16
            for key, value in gantt_data.performance_metrics.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_tasks_sheet(self, wb: openpyxl.Workbook, gantt_data: ConstellationGanttData):
        """创建任务详情工作表"""
        ws = wb.create_sheet("任务详情")
        
        # 准备数据
        tasks_data = []
        for task in gantt_data.tasks:
            threat_level = getattr(task, 'threat_level', 3)
            priority = getattr(task, 'priority', 3)
            quality_score = getattr(task, 'quality_score', 0.8)
            
            tasks_data.append({
                '任务ID': task.task_id,
                '任务名称': task.task_name,
                '分配卫星': task.assigned_satellite,
                '目标导弹': task.target_missile,
                '任务类别': task.category,
                '威胁等级': threat_level,
                '优先级': priority,
                '执行状态': task.execution_status,
                '质量分数': quality_score,
                '开始时间': task.start_time,
                '结束时间': task.end_time,
                '持续时间(分钟)': (task.end_time - task.start_time).total_seconds() / 60
            })
        
        # 创建DataFrame并写入Excel
        df = pd.DataFrame(tasks_data)
        
        # 写入表头
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # 写入数据
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                
                # 根据威胁等级设置颜色
                if col_num == 6:  # 威胁等级列
                    threat_level = int(value) if isinstance(value, (int, float)) else 3
                    color = self.threat_colors.get(threat_level, self.threat_colors[3])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(color='FFFFFFFF', bold=True)
                
                # 根据状态设置颜色
                elif col_num == 8:  # 执行状态列
                    status = str(value).lower()
                    color = self.status_colors.get(status, self.status_colors['planned'])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(color='FFFFFFFF', bold=True)
        
        # 设置列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_satellites_sheet(self, wb: openpyxl.Workbook, gantt_data: ConstellationGanttData):
        """创建卫星信息工作表"""
        ws = wb.create_sheet("卫星信息")
        
        # 统计每颗卫星的任务
        satellite_stats = {}
        for satellite in gantt_data.satellites:
            satellite_tasks = [t for t in gantt_data.tasks if t.assigned_satellite == satellite]
            
            if satellite_tasks:
                total_duration = sum((t.end_time - t.start_time).total_seconds() for t in satellite_tasks)
                avg_threat = sum(getattr(t, 'threat_level', 3) for t in satellite_tasks) / len(satellite_tasks)
                
                satellite_stats[satellite] = {
                    '卫星ID': satellite,
                    '任务数量': len(satellite_tasks),
                    '总工作时长(小时)': total_duration / 3600,
                    '平均威胁等级': avg_threat,
                    '工作负载': total_duration / (gantt_data.end_time - gantt_data.start_time).total_seconds()
                }
        
        # 写入数据
        headers = ['卫星ID', '任务数量', '总工作时长(小时)', '平均威胁等级', '工作负载']
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
        
        # 写入数据
        for row_num, (satellite, stats) in enumerate(satellite_stats.items(), 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = stats[header]
                
                if isinstance(value, float):
                    cell.value = round(value, 2)
                else:
                    cell.value = value
        
        # 设置列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
    
    def _create_timeline_sheet(self, wb: openpyxl.Workbook, gantt_data: ConstellationGanttData):
        """创建时间线工作表"""
        ws = wb.create_sheet("时间线")
        
        # 创建简化的甘特图表格
        ws['A1'] = "卫星"
        ws['B1'] = "任务"
        ws['C1'] = "开始时间"
        ws['D1'] = "结束时间"
        ws['E1'] = "持续时间"
        ws['F1'] = "威胁等级"
        
        # 设置表头样式
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
        
        # 按开始时间排序任务
        sorted_tasks = sorted(gantt_data.tasks, key=lambda t: t.start_time)
        
        # 写入任务数据
        for row_num, task in enumerate(sorted_tasks, 2):
            ws[f'A{row_num}'] = task.assigned_satellite
            ws[f'B{row_num}'] = task.task_name
            ws[f'C{row_num}'] = task.start_time
            ws[f'D{row_num}'] = task.end_time
            ws[f'E{row_num}'] = (task.end_time - task.start_time).total_seconds() / 60
            ws[f'F{row_num}'] = getattr(task, 'threat_level', 3)
            
            # 设置威胁等级颜色
            threat_cell = ws[f'F{row_num}']
            threat_level = getattr(task, 'threat_level', 3)
            color = self.threat_colors.get(threat_level, self.threat_colors[3])
            threat_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            threat_cell.font = Font(color='FFFFFFFF', bold=True)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
    
    def _create_statistics_sheet(self, wb: openpyxl.Workbook, gantt_data: ConstellationGanttData):
        """创建统计分析工作表"""
        ws = wb.create_sheet("统计分析")
        
        # 威胁等级分布
        threat_distribution = {}
        for task in gantt_data.tasks:
            threat_level = getattr(task, 'threat_level', 3)
            threat_distribution[threat_level] = threat_distribution.get(threat_level, 0) + 1
        
        ws['A1'] = "威胁等级分布"
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A3'] = "威胁等级"
        ws['B3'] = "任务数量"
        
        row = 4
        for level, count in sorted(threat_distribution.items()):
            ws[f'A{row}'] = f"等级 {level}"
            ws[f'B{row}'] = count
            
            # 设置颜色
            color = self.threat_colors.get(level, self.threat_colors[3])
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[f'A{row}'].font = Font(color='FFFFFFFF', bold=True)
            
            row += 1
        
        # 状态分布
        status_distribution = {}
        for task in gantt_data.tasks:
            status = task.execution_status
            status_distribution[status] = status_distribution.get(status, 0) + 1
        
        ws['A10'] = "执行状态分布"
        ws['A10'].font = Font(size=14, bold=True)
        
        ws['A12'] = "执行状态"
        ws['B12'] = "任务数量"
        
        row = 13
        for status, count in status_distribution.items():
            ws[f'A{row}'] = status
            ws[f'B{row}'] = count
            
            # 设置颜色
            color = self.status_colors.get(status, self.status_colors['planned'])
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[f'A{row}'].font = Font(color='FFFFFFFF', bold=True)
            
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

# 全局Excel生成器实例
_excel_gantt_generator = None

def get_gantt_excel_generator() -> AerospaceGanttExcelGenerator:
    """获取全局甘特图Excel生成器实例"""
    global _excel_gantt_generator
    if _excel_gantt_generator is None:
        _excel_gantt_generator = AerospaceGanttExcelGenerator()
    return _excel_gantt_generator
